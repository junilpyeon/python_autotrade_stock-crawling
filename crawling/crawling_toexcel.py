import urllib.request
from bs4 import BeautifulSoup
from openpyxl import Workbook


def main():
    url = "http://www.newsis.com/realnews/"
    sourcecode = urllib.request.urlopen(url).read()
    soup = BeautifulSoup(sourcecode, "html.parser")

    articles = []

    for i in soup.find_all("strong", class_="title"):
        articles.append(i.get_text())

    wb = Workbook()
    sheet1 = wb.active
    file_name = 'output.xlsx'

    for i in range(0, len(articles)):
        sheet1.cell(row=i + 1, column=1).value = i + 1
        sheet1.cell(row=i + 1, column=2).value = articles[i]

    wb.save(filename=file_name)


if __name__ == "__main__":
    main()
